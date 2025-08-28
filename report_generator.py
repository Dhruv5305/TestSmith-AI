import json
import os
from datetime import datetime
from config import REPORTS_DIR
import openpyxl
from openpyxl import Workbook
import logging

logger = logging.getLogger('ReportGenerator')

def add_to_report(report_entry):
    """
    Add a test result to the JSON report
    """
    # Ensure reports directory exists
    os.makedirs(REPORTS_DIR, exist_ok=True)
    
    report_file = os.path.join(REPORTS_DIR, 'test_results.json')
    
    # Load existing report or create new one
    if os.path.exists(report_file):
        with open(report_file, 'r') as f:
            try:
                report_data = json.load(f)
            except json.JSONDecodeError:
                report_data = {"results": []}
    else:
        report_data = {"results": []}
    
    # Add new entry
    report_data["results"].append(report_entry)
    
    # Save updated report
    with open(report_file, 'w') as f:
        json.dump(report_data, f, indent=2)
    
    logger.info(f"Added test result to report: {report_entry['test_name']} - {report_entry['status']}")

def generate_excel_report(results=None):
    """
    Generate an Excel report from the JSON results
    """
    if results is None:
        # Load from JSON file
        report_file = os.path.join(REPORTS_DIR, 'test_results.json')
        if not os.path.exists(report_file):
            logger.error("No JSON report found to convert to Excel")
            return
        
        with open(report_file, 'r') as f:
            report_data = json.load(f)
        results = report_data.get("results", [])
    
    if not results:
        logger.warning("No results to generate Excel report")
        return
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    # Add headers
    headers = ["Test Name", "Description", "Status", "Error", "Screenshot", "Generated Code", "Timestamp"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add data
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result.get('test_name', ''))
        ws.cell(row=row, column=2, value=result.get('description', ''))
        ws.cell(row=row, column=3, value=result.get('status', ''))
        ws.cell(row=row, column=4, value=result.get('error', ''))
        ws.cell(row=row, column=5, value=result.get('screenshot_link', ''))
        ws.cell(row=row, column=6, value=result.get('generated_code_link', ''))
        ws.cell(row=row, column=7, value=result.get('timestamp', ''))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_file = os.path.join(REPORTS_DIR, f'test_results_{timestamp}.xlsx')
    wb.save(excel_file)
    
    logger.info(f"Excel report generated: {excel_file}")
    return excel_file